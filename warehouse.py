from openpyxl import load_workbook, Workbook
import pandas as pd
from support_module import email_automation
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')
import glob
from datetime import datetime
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ===== Parameters =====


# 信箱下載 DSV 進櫃計畫表 相關設置
SUBJECT_KEYWORD = "進櫃通知單" #找出含有此關鍵字的最新email並下載他的附件
SAVE_DIR = r"YOUR_NETWORK_DRIVE\ShipmentTracking\DSV進櫃計畫表" #下載DSV進櫃計畫表的資料夾
FOLDER_PATH = "收件匣" # 英文 Outlook 用 "Inbox"；或直接用預設 GetDefaultFolder(6) 不理它


#DSV 進櫃通知單 前處理
DSV_SHEET = "DSV-進櫃計畫表" #DSV進櫃計畫表裡面要讀取的工作表名稱
HEADER_ROW = 3 #這張表的標題列在第三列
MAX_SCAN_ROW = 500 #這張表因人工操作不佳使列數到很遠，需要設置邊界否則跑不完
MAX_SCAN_COL = 50 #這張表因人工操作不佳使欄位到很遠，需要設置邊界否則跑不完
TARGET_HEADERS = ["Container No", "預計到櫃時間"] #解開合併儲存格後要往下填充的目標欄位名稱


#KAO 相關設置
KAO_DIR = r"YOUR_NETWORK_DRIVE\ShipmentTracking\SalesData\Daily Report" #KAO報表所在資料夾路徑
KAO_FILENAME_KEYWORD = "KAO" #找出KAO_DIR資料夾中含有此關鍵字的最新的xlsx檔案
KAO_SHEET = "可提櫃" #KAO報表裡面要讀取的工作表名稱


# ======================

def get_paths():
    #收件匣裡面找最新的符合條件的附件，下載到指定資料夾，回傳檔案路徑列表
    files = email_automation.download_latest_matching_attachments(
    subject_contains=SUBJECT_KEYWORD,      #郵件主旨包含這個字串才會被下載
    save_dir=SAVE_DIR,
    mailbox=None,                     # 若有多信箱才需要填
    folder_path=FOLDER_PATH,             # 英文 Outlook 用 "Inbox"；或直接用預設 GetDefaultFolder(6) 不理它
    allowed_ext=(".xlsx",),                 # 例：(".pdf", ".xlsx")
    max_items_scan=300
    )
    dsv_path = files[0] if files else None
    dsv_sheet = DSV_SHEET
    # Dynamically find the latest KAO xlsx file
    kao_dir = KAO_DIR
    kao_files = glob.glob(os.path.join(kao_dir, f'*{KAO_FILENAME_KEYWORD}*.xlsx'))
    if not kao_files:
        raise FileNotFoundError(f"No KAO xlsx files found in {kao_dir}")
    kao_path = max(kao_files, key=os.path.getmtime)
    kao_sheet = KAO_SHEET

    print(f"Loading DSV file: {dsv_path}, sheet: {dsv_sheet}")
    print(f"Loading KAO file: {kao_path}, sheet: {kao_sheet}")
    #for test
    #dsv_path = r"YOUR_NETWORK_DRIVE\ShipmentTracking\DSV進櫃計畫表\DSV進櫃計劃表-20260226_new2.xlsx" #可被email取代

    return dsv_path, dsv_sheet, kao_path, kao_sheet

#決定如何選擇進倉時間：如果廣吉或CEVA其中一個有值且以"X"結尾，就用另一個；如果兩個都有值但都不以"X"結尾，就用第一個有數字的；如果都沒有數字就不改動該列
def get_inbound_time(row, kao_df):
    material = row['Material']
    po_no = row['Purchasing Document']
    match = kao_df[
        (kao_df['Item Code'] == material) &
        ((kao_df['Po. 1'] == po_no) | (kao_df['Po. 2'] == po_no))
    ]
    if not match.empty:
        print(f"Found KAO match for Material={material}, PO No={po_no}")
        kao_row = match.iloc[0]
        guangji_time = str(kao_row['廣吉進櫃時間'])
        ceva_time = str(kao_row['CEVA進櫃時間'])
        print(f"廣吉進櫃時間: {guangji_time}, CEVA進櫃時間: {ceva_time}")
        # Rule: If one ends with "X", use the other
        if guangji_time.endswith("X"):
            print(f"廣吉進櫃時間 ends with 'X', using CEVA time: {ceva_time}")
            return ceva_time
        elif ceva_time.endswith("X"):
            print(f"CEVA進櫃時間 ends with 'X', using 廣吉 time: {guangji_time}")
            return guangji_time
        # If neither ends with "X", pick the first with a numeric value
        for t in [guangji_time, ceva_time]:
            if any(char.isdigit() for char in t):
                print(f"Using numeric time: {t}")
                return t
    return row['進倉時間']  # fallback to existing value

def DSV(dsv_path, dsv_sheet):
        # 讀原始檔
    wb_src = load_workbook(dsv_path, data_only=True) #原表有會造成開不起來的公式
    ws_src = wb_src[dsv_sheet]

    # 建立新檔
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = dsv_sheet

    max_row = min(ws_src.max_row, MAX_SCAN_ROW)
    max_col = MAX_SCAN_COL

    # 1) 先把整張 sheet 的值複製到新檔（只複製值，不複製格式），複製到新檔案時，合併儲存格會自動被拆開，這時候只要把值向下填充即可
    for r in range(HEADER_ROW, max_row + 1): #從第三列標題列開始
        row_values = [ws_src.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        ws_new.append(row_values)

    print(f"Copied values from source to new sheet. Total rows: {max_row}, columns: {max_col}")
    for r in range(1, 3):  # Rows 1 and 2 (1-based index)
        row_values = [ws_new.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        #print(f"Row {r}: {row_values}")


    # 1. 刪除多餘列("抵達臺灣日"為空)，目前都找不到
    arrive_col_idx = None
    for c in range(1, max_col + 1):
        header = ws_new.cell(row=1, column=c).value
        if isinstance(header, str) and header.strip() == "抵達臺灣日":
            arrive_col_idx = c
            break

    if arrive_col_idx is not None:
        # 2. Collect rows to delete (where "抵達臺灣日" is empty)
        rows_to_delete = []
        for row in range(2, ws_new.max_row + 1):
            cell_value = ws_new.cell(row=row, column=arrive_col_idx).value
            if cell_value is None or str(cell_value).strip() == "":
                rows_to_delete.append(row)
        # 3. Delete rows from bottom to top
        for row in reversed(rows_to_delete):
            ws_new.delete_rows(row)
    else:
        print("找不到 '抵達臺灣日' 欄位，無法刪除空值列")

    # 2) 找前 50 欄中的目標欄位(container no 和 預計到櫃時間)的欄位索引，記錄在 col_map 裡面，並往下填充
    col_map = {}
    scan_limit = min(MAX_SCAN_COL, max_col)

    for c in range(1, scan_limit + 1):
        header = ws_src.cell(row=HEADER_ROW, column=c).value
        #print(f"Scanning column {c} for header: {header}")
        if isinstance(header, str):
            header = header.strip()
        if header in TARGET_HEADERS:
            col_map[header] = c

    print("Target column indices:", col_map)

    missing = [h for h in TARGET_HEADERS if h not in col_map]
    if missing:
        raise ValueError(f"找不到目標欄位: {missing}")
    # Fill down the target columns in ws_new
    for header, col_idx in col_map.items():
        last_value = None
        for row in range(HEADER_ROW + 1, max_row + 1):  # Start after header row
            cell = ws_new.cell(row=row, column=col_idx)
            if cell.value is None:
                cell.value = last_value
            else:
                last_value = cell.value

    output_path = os.path.join(os.path.dirname(dsv_path), os.path.basename(dsv_path).split('.')[0] + '_clean.xlsx')
    wb_new.save(output_path)
    print(f"Done. Saved to: {output_path}")


    ##output_path = r"YOUR_NETWORK_DRIVE\ShipmentTracking\DSV進櫃計劃表\DSV進櫃計劃表-20260320__20260320_182540_clean.xlsx" #for test
    dsv_df = pd.read_excel(output_path)
    print(len(dsv_df))

    # Set [報關行] = "V" for all rows with the same [Container No] if any row has [報關行] == "V"
    processed_containers = set()
    for idx, row in dsv_df.iterrows():
        if row['報關行'] == "V":
            container_no = row['Container No']
            if container_no not in processed_containers:
                dsv_df.loc[dsv_df['Container No'] == container_no, '報關行'] = "V"
                processed_containers.add(container_no)

    #print(dsv_df['Po.No'].tail(5))
    #print(dsv_df['Item Code'].tail(5))
    print("Processed DSV dates and V updates.")
    return dsv_df


def warehouse(po_update_path):
    print("Starting warehouse module...")
    dsv_path, dsv_sheet, kao_path, kao_sheet = get_paths()

    ## ---------- DSV ---------- ##
    dsv_df = DSV(dsv_path, dsv_sheet)


    #recalculate 重算一次讓值全部進入快取裡，待會才能使用Cargoo公式產出的資料
    #support_function.recalculate_workbook(po_update_path)

    wb_po = load_workbook(po_update_path)
    ws_po = wb_po["大表"]

    # Read PO sheet into pandas
    po_df = pd.read_excel(po_update_path, sheet_name="大表", header=1)  # Assuming header is in the second row (index 1)
    po_df['保稅進倉時間'] = po_df['保稅進倉時間'].astype('object')
    po_df['進倉時間'] = po_df['進倉時間'].astype('object')

    # 4. Update PO pandas DataFrame row by row
    for idx, po_row in po_df.iterrows():
        material = po_row['Material']
        po_no = po_row['Purchasing Document']
        match = dsv_df[(dsv_df['Item Code'] == material) & (dsv_df['Po.No'] == po_no)].iloc[::-1]
        #print(f"Processing PO row {idx}: Material={material}, PO No={po_no}")
        if not match.empty:
            print(f"Processing PO row {idx}: Material={material}, PO No={po_no}, Matches found: {len(match)}")
            dsv_row = match.iloc[0]
            if dsv_row['報關行'] == "V" and "保稅提回" not in str(dsv_row['預計到櫃時間']):
                po_df.at[idx, '保稅進倉時間'] = dsv_row['預計到櫃時間']
                po_df.at[idx, '進倉時間'] = "-"
            else:
                po_df.at[idx, '保稅進倉時間'] = "-"
                po_df.at[idx, '進倉時間'] = dsv_row['預計到櫃時間']
        else:
            po_df.at[idx, '保稅進倉時間'] = "-"
            po_df.at[idx, '進倉時間'] = "-"

    print("Updated PO DataFrame with DSV information.") 



    ## ---------- KAO ---------- ##
    # Read KAO sheet
    kao_df = pd.read_excel(kao_path, sheet_name=kao_sheet)
    print(f"Loaded KAO data. Total rows: {len(kao_df)}")

    # Filter by [新增日] for this year and last year
    today = pd.Timestamp.today()
    current_year = today.year
    last_year = current_year - 1
    kao_df['新增日'] = pd.to_datetime(kao_df['新增日'], errors='coerce')
    kao_df = kao_df[kao_df['新增日'].dt.year.isin([current_year, last_year])]

    # Update [進倉時間] in PO DataFrame based on KAO rules
    for idx, po_row in po_df.iterrows():
        inbound_time = get_inbound_time(po_row, kao_df)
        po_df.at[idx, '進倉時間'] = inbound_time
    print("Updated PO DataFrame with KAO information.")


    # 5. ---------Write back only the updated columns to preserve formatting---------
    material_col = None
    po_no_col = None
    baoshui_col = None
    jin_cang_col = None

    # Find column indices #都在第二列
    for col in ws_po.iter_cols(1, ws_po.max_column):
        header = col[1].value
        if header == 'Material':
            material_col = col[1].column
        elif header == 'Purchasing Document':
            po_no_col = col[1].column
        elif header == '保稅進倉時間':
            baoshui_col = col[1].column
        elif header == '進倉時間':
            jin_cang_col = col[1].column

    # Update values row by row，寫入原始PO_update資料
    for i in range(len(po_df)):
        po_row = po_df.iloc[i]
        excel_row = i + 3  # If your data starts at row 3 (adjust if needed)
        if baoshui_col:
            ws_po.cell(row=excel_row, column=baoshui_col, value=po_row['保稅進倉時間'])
        if jin_cang_col:
            ws_po.cell(row=excel_row, column=jin_cang_col, value=po_row['進倉時間'])

    wb_po.save(po_update_path)
    print("PO update completed and formatting preserved.")
    print("finished warehouse module.")
    print("-" * 50)



if __name__ == "__main__":
    po_update_path = r"YOUR_NETWORK_DRIVE\ShipmentTracking\PO update\temp\PO update 20260225_experiment_1.xlsx"
    kao_path = r"YOUR_NETWORK_DRIVE\ShipmentTracking\DSV進櫃計畫表\0319 KAO_test_only.xlsx"
    warehouse(po_update_path)