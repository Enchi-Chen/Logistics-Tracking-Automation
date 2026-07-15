import os
import re
import glob
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


#主流程在最後一個函式
SRC_DIR = r"YOUR_NETWORK_DRIVE\ShipmentTracking\SalesData\Daily Report" #報關行報表所在資料夾路徑
OUTPUT_PATH = r"YOUR_NETWORK_DRIVE\ShipmentTracking\customs_combined.xlsx" #裝合併後的檔案路徑
MB_SHEETS = ["放行", "報關中", "未報關"]
GENERAL_SHEETS = ["可提櫃", "報關中", "未報關"]
#除了"放行"會被改為"可提櫃"，其他的工作表名稱都會被直接寫入大表。
#若有新的工作表名稱，請放到上方的MB_SHEETS或GENERAL_SHEETS清單中，程式會自動抓取該工作表內容並寫入大表。


def find_col(df, patterns):
    """Return first column name that matches any regex pattern in patterns (case-insensitive)."""
    for col in df.columns:
        for p in patterns:
            if re.search(p, str(col), flags=re.I): #flags是什麼?
                return col
    return None

def map_po_columns(df, rule=None):
    """Find two PO columns (PO1, PO2).
    If rule == 'MB' (MB file) pick Po. 1 as PO1 and Po. 3 as PO2 (exclude Po. 2).
    Otherwise fall back to generic detection.
    """
    if rule == "MB":
        po1 = None
        po2 = None
        # try explicit patterns for Po. 1 and Po. 3 (allow newline/extra chars)
        for c in df.columns:
            if po1 is None and re.search(r"Po[\.\s\W]*1\b", str(c), flags=re.I):
                po1 = c
            if po2 is None and re.search(r"Po[\.\s\W]*3\b", str(c), flags=re.I):
                po2 = c
            if po1 and po2:
                break
        # fallback to generic if not found
        if po1 is None or po2 is None:
            """Find two PO columns (PO1, PO2) by matching column names containing 'Po' or 'PO'."""
            po_cols = [c for c in df.columns if re.search(r"\bPo\b|^Po\.|^PO\b|Po\s*\.", str(c), flags=re.I)]
            if not po_cols:
                po_cols = [c for c in df.columns if re.search(r"Po", str(c), flags=re.I)]
            if po1 is None and po_cols:
                po1 = po_cols[0]
            if po2 is None and len(po_cols) >= 2:
                po2 = po_cols[1]
        return po1, po2

    # default/generic behavior
    po_cols = [c for c in df.columns if re.search(r"\bPo\b|^Po\.|^PO\b|Po\s*\.", str(c), flags=re.I)]
    if not po_cols:
        po_cols = [c for c in df.columns if re.search(r"Po", str(c), flags=re.I)]
    po1 = po_cols[0] if len(po_cols) >= 1 else None
    po2 = po_cols[1] if len(po_cols) >= 2 else None
    return po1, po2


def process_file(path):
    base = os.path.basename(path)
    print("base: ", base)
    name_upper = base.upper()
    print("name_upper: ", name_upper)
    records = []
    # Determine rule / sheet names and header row
    is_MB = ("MB" in name_upper or "NHS" in name_upper)
    is_general = any(k in name_upper for k in ("GENERAL", "KAO", "NPP", "REEFER"))

    if is_MB:
        sheets = MB_SHEETS
        header_row = 1  # use second row as header
    elif is_general:
        sheets = GENERAL_SHEETS
        header_row = 0
    else:
        return []  # skip files not matching rules

    # cutoff = today - 6 months
    cutoff = pd.Timestamp.today().normalize() - pd.DateOffset(months=6)

    for sht in sheets:
        try:
            df = pd.read_excel(path, sheet_name=sht, header=header_row, engine="openpyxl")
        except Exception:
            continue

        # Normalize columns by stripping whitespace from column names
        df.columns = [str(c).strip() for c in df.columns]

        # Find Item Code column
        item_col = find_col(df, [r"\bItem Code\b", r"Item\s*Code", r"ItemCode"])
        # PO columns
        po1_col, po2_col = map_po_columns(df, rule="MB" if is_MB else None)
        # Container No
        container_col = find_col(df, [r"Container\s*No", r"Container"])
        # Free Time
        free_col = find_col(df, [r"Free\s*Time", r"FreeTime"])
        # Notice
        notice_col = find_col(df, [r"Notice*"])
        # Date
        date_col = find_col(df, [r"新增日", r"新增 日"])
        # ETA
        eta_col = find_col(df, [r"^ETA.*"])

        # date filtering
            # propagate merged-cell values down the column
        if "NPP" in name_upper:
            df[date_col] = df[date_col].ffill()
            # convert to datetime (invalid -> NaT)
        
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
        # keep only rows with non-empty date and date > cutoff
        df = df[df[date_col].notna() & (df[date_col] > cutoff)]
        # if df.empty:
        #     continue

        # Build resulting frame with selected columns
        out = pd.DataFrame()
        out["Item Code"] = df[item_col] if item_col in df.columns else pd.NA
        out["PO1"] = df[po1_col] if po1_col in df.columns else pd.NA
        out["PO2"] = df[po2_col] if po2_col in df.columns else pd.NA
        out["Status"] = "可提櫃" if sht == "放行" else sht # Set Status: if sht is '放行', use '可提櫃', else use sht
        out["Container No"] = df[container_col] if container_col in df.columns else pd.NA
        out["Free Time"] = df[free_col] if free_col in df.columns else pd.NA
        out["Notice"] = df[notice_col] if notice_col in df.columns else pd.NA
        # Origin is file name (without path)
        out["Origin"] = base
        out["Date"] = df[date_col] if date_col in df.columns else pd.NA
        out["ETA"] = df[eta_col] if eta_col in df.columns else pd.NA

        records.append(out)

    return records


# Item Code cleansing:
# - If parentheses present, keep the value inside the first parentheses.
# - If contains 'TBC' and a number, keep only the number. If it is 'TBC' alone, leave as 'TBC'.
def _clean_item(val):
    if pd.isna(val):
        return pd.NA
    s = str(val).strip()

    # remove trailing ".0" (common when Excel floats become strings)
    if s.endswith('.0'):
        s = s[:-2].strip()

    # parentheses
    m = re.search(r'\(([^)]*)\)', s)
    if m:
        inner = m.group(1).strip()
        return inner if inner != '' else pd.NA
    return s

# - Convert to string and remove any word that starts with 'LDP' (case-insensitive).
def _clean_po(val):
    if pd.isna(val):
        return pd.NA
    s = str(val)
    # remove words starting with LDP
    s2 = re.sub(r'(?i)\bLDP\w*\b', '', s)
    # collapse whitespace and trim punctuation leftover
    s2 = re.sub(r'[\s\-\>\_\,\;\:]+', ' ', s2).strip()

    # remove trailing ".0" (common when Excel floats become strings)
    if s2.endswith('.0'):
        s2 = s2[:-2].strip()

    # TBC with numbers
    if re.search(r'(?i)\bTBC', s2):
        m = re.search(r'(\d+)', s2)
        if m:
            return m.group(1)
        return 'TBC'

    return s2 if s2 != '' else pd.NA

# Clean ETA: strip trailing "00:00:00" (optional leading space or 'T'), return pd.NA for empty/nan
def _clean_eta(val):
    if pd.isna(val):
        return pd.NA
    s = str(val).strip()
    # remove trailing time "00:00:00" possibly preceded by space or 'T'
    s = re.sub(r'(?:[\sT]*00:00:00)$', '', s)
    s = s.strip()
    if s.lower() in ('nan', 'nat', ''):
        return pd.NA
    return s

def combine_all():
    #至YOUR_NETWORK_DRIVE\ShipmentTracking\SalesData\Daily Report，擷取所有位於該位置的Excel檔案，個別抓取正確欄位後，合併成combined
    files = glob.glob(os.path.join(SRC_DIR, "*.xlsx*"))
    parts = []
    for f in reversed(files):
        parts.extend(process_file(f))

    if not parts:
        print("No matching files/sheets found.")
        return pd.DataFrame(columns=["Item Code","PO1","PO2","Status","Container No","Free Time","Origin","Date","Notice","ETA"])

    combined = pd.concat(parts, ignore_index=True)
    
    #將每個欄位的資料型態進行前處理，包含統一為字串、去除前後空白、將nan轉為pd.NA，以及針對Item Code、PO1、PO2、ETA欄位進行特定的清理規則
    # Ensure string type and normalize NaN
    for c in ["Item Code","PO1","PO2","Container No","Free Time","Origin","Date","Notice","ETA"]:
        if c in combined.columns:
            combined[c] = combined[c].astype(str).str.strip().replace({"nan": pd.NA})
    #item code cleansing
    if "Item Code" in combined.columns:
        combined["Item Code"] = combined["Item Code"].apply(_clean_item)
    # PO1 cleansing:
    if "PO1" in combined.columns:
        combined["PO1"] = combined["PO1"].apply(_clean_po)
    if "PO2" in combined.columns:
        combined["PO2"] = combined["PO2"].apply(_clean_po)
    # ETA cleansing:
    if "ETA" in combined.columns:
        combined["ETA"] = combined["ETA"].apply(_clean_eta)

    combined.to_excel(OUTPUT_PATH, index=False)
    print(f"Saved combined file to: {OUTPUT_PATH}")
    return combined



def write_to_target(po_update_path,combined):

    # prepare combined lookup maps (strings, trimmed)
    lookup = combined.copy()
    for c in ["PO1", "PO2", "Item Code", "Notice", "Free Time", "Status"]:
        if c in lookup.columns:
            lookup[c] = lookup[c].astype(str).str.strip().replace({"nan": None})

    # build PO -> first-row-index map (PO1 preferred, then PO2 if PO1 missing)
    po_map = {}
    if "PO1" in lookup.columns:
        for idx, v in lookup["PO1"].dropna().items():
            if v and v not in po_map:
                po_map[v] = idx
    if "PO2" in lookup.columns:
        for idx, v in lookup["PO2"].dropna().items():
            if v and v not in po_map:
                po_map[v] = idx

    # build (PO, Item) -> first-row-index map (check PO1 first then PO2)
    po_item_map = {}
    if "PO1" in lookup.columns and "Item Code" in lookup.columns:
        for idx, row in lookup.iterrows():
            po = row.get("PO1")
            it = row.get("Item Code")
            if po and it:
                key = (po, it)
                if key not in po_item_map:
                    po_item_map[key] = idx
    if "PO2" in lookup.columns and "Item Code" in lookup.columns:
        for idx, row in lookup.iterrows():
            po = row.get("PO2")
            it = row.get("Item Code")
            if po and it:
                key = (po, it)
                if key not in po_item_map:
                    po_item_map[key] = idx

    # load workbook and worksheet (preserve formulas/format)
    wb = load_workbook(po_update_path)
    sheet_name = "大表" if "大表" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    header_row = 2  # header is on second row

    # helper: find header column index by regex patterns (case-insensitive)
    def find_header_col(patterns):
        import re
        for cell in ws[header_row]:
            val = cell.value
            if not val:
                continue
            s = str(val).strip()
            for p in patterns:
                if re.search(p, s, flags=re.I):
                    return cell.col_idx
        return None

    def get_or_create_col(header_name):
        # try exact match first (strip & lower)
        for cell in ws[header_row]:
            if cell.value and str(cell.value).strip().lower() == header_name.strip().lower():
                return cell.col_idx
        # not found -> append as new column
        new_col = ws.max_column + 1
        ws.cell(row=header_row, column=new_col, value=header_name)
        return new_col

    # locate purchasing & material columns
    purch_col = find_header_col([r"Purchasing Document", r"Purchase Document", r"Purchasing.*Doc", r"\bPO\b"])
    mat_col = find_header_col([r"Material", r"Material Number"])

    if purch_col is None and mat_col is None:
        print("Cannot find Purchasing or Material columns in update file headers.")
        exit(1)
        #return

    # ensure target columns exist (will preserve other cells)
    notice_po_col = get_or_create_col("Notice(first PO)")
    notice_po_item_col = get_or_create_col("Notice(first PO+item)")
    free_time_col = get_or_create_col("Free Time")
    status_col = get_or_create_col("Status")
    eta_col = get_or_create_col("ETA")


    #找最後一列(以PO, Material至少其一為非空值為準)
    def find_last_data_row(ws, start_row=1, end_row=None, cols=None):
        """Return last row index that has any non-empty cell.value.
        cols: iterable of 1-based column indexes to check; if None check all columns."""
        if end_row is None:
            end_row = ws.max_row
        cols_to_check = cols if cols is not None else range(1, ws.max_column + 1)
        for r in range(end_row, start_row - 1, -1):
            for c in cols_to_check:
                val = ws.cell(row=r, column=c).value
                if val not in (None, ''):
                    return r
        return start_row - 1

    # use only purchasing/material columns if available to speed up scan
    cols_to_scan = None
    if purch_col or mat_col:
        cols_to_scan = [c for c in (purch_col, mat_col) if c]

    upper_bound = ws.max_row
    last_row = find_last_data_row(ws, start_row=header_row + 1, end_row=upper_bound, cols=cols_to_scan)

    # iterate data rows (rows below header_row)，開始尋找與寫入資料
    for r in range(header_row + 1, last_row + 1):
        purch_val = None
        mat_val = None

        if purch_col:
            purch_val = ws.cell(row=r, column=purch_col).value
            purch_val = str(purch_val).strip() if purch_val is not None else ""
        if mat_col:
            mat_val = ws.cell(row=r, column=mat_col).value
            mat_val = str(mat_val).strip() if mat_val is not None else ""

        # if r > 390 and r < 394:
        #     print(purch_val)
        #     print(mat_val)
        
        # Notice(first PO): find by PO (PO1 first then PO2 via our po_map)
        notice_val = "-"
        if purch_val:
            idx_po = po_map.get(purch_val)
            if idx_po is not None:
                notice_lookup = lookup.at[idx_po, "Notice"] if "Notice" in lookup.columns else None
                if notice_lookup in (None, "", "None"):
                    notice_val = "<NA>"
                else:
                    notice_val = notice_lookup
        ws.cell(row=r, column=notice_po_col, value=notice_val)

        # Notice(first PO+item), Free Time, Status, ETA: find by (PO, Item)
        notice_po_item_val = "-"
        free_time_val = "-"
        status_val = "-"
        eta_val = "-"
        if purch_val and mat_val:
            idx_po_item = po_item_map.get((purch_val, mat_val))
            if idx_po_item is not None:
                notice_po_item_val = lookup.at[idx_po_item, "Notice"] if "Notice" in lookup.columns else None
                if notice_po_item_val in (None, "", "None","<NA>"):
                    notice_po_item_val = "<NA>"
                free_time_val = lookup.at[idx_po_item, "Free Time"] if "Free Time" in lookup.columns else None
                if free_time_val in (None, "", "None","<NA>"):
                    free_time_val = "<NA>"
                status_val = lookup.at[idx_po_item, "Status"] if "Status" in lookup.columns else None
                if status_val in (None, "", "None"):
                    status_val = "<NA>"
                eta_val = lookup.at[idx_po_item, "ETA"] if "ETA" in lookup.columns else None
                if eta_val in (None, "", "None"):
                    eta_val = "<NA>"


        ws.cell(row=r, column=notice_po_item_col, value=notice_po_item_val)
        # write free time and status (keep blank if None)
        ws.cell(row=r, column=free_time_col, value=free_time_val if free_time_val not in (None, "None") else None)
        ws.cell(row=r, column=status_col, value=status_val if status_val not in (None, "None") else None)
        ws.cell(row=r, column=eta_col, value=eta_val if eta_val not in (None, "None") else None)

    # save workbook (in-place) — preserves other sheets, formulas, formats
    wb.save(po_update_path)
    print("Updated file saved:", po_update_path)
    # call modifier before writing combined to OUTPUT_PATH


def customs(po_update_path):
    print("Starting customs module...")
    combined = combine_all()
    write_to_target(po_update_path, combined)
    print("Customs module finished, combined data written to update file.")
    print("-" * 50)


