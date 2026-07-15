
import asyncio
import time
from math import ceil
import os
import sys
#from io import StringIO
from support_module import support_function

#-------default programs----------#


from carrier_tracking import Evergreen_play
from carrier_tracking import YangMing_API
from carrier_tracking import Maersk_API
from carrier_tracking import ONE_API
from carrier_tracking import WanHai_play

#-------backup programs----------#

#import YangMing #selenium
#from carrier_tracking import Evergreen_API ##經測試，資料不齊全
#import Maersk #selenium
#import ONE #selenium
#import WanHai #Selenium
#import OOCL


import pandas as pd
from openpyxl import load_workbook

import time 

#選擇要執行的carrier
CARRIER_LIST = ["Evergreen","Wan Hai", "Yang Ming Line","Maersk Line","Ocean Network Express"] #"Evergreen","Wan Hai", "Yang Ming Line","Maersk Line","Ocean Network Express"
FILTERED_DF_OUTPUT_PATH = r"YOUR_NETWORK_DRIVE\ShipmentTracking\carrier_tracking_main_temp\filtered_df_output.csv" #儲存過濾後的Carrier和Booking資訊，供carrier tracking模組使用
RESULT_DF_OUTPUT_PATH = r"YOUR_NETWORK_DRIVE\ShipmentTracking\carrier_tracking_main_temp\final_data.csv" #儲存carrier tracking模組的最終結果



def track(df):
    track_start = time.time()
    try:
        # Convert to a DataFrame
        df = df.rename(columns={"Carrier": "Carrier Name", 'Booking': 'Booking Number'})
        
        # Filter booking numbers for companies, all components should be string type for later comparison
        evergreen_bookings = list(map(str, df[df["Carrier Name"] == "Evergreen"]["Booking Number"].tolist()))
        wan_hai_bookings = list(map(str, df[df["Carrier Name"] == "Wan Hai"]["Booking Number"].tolist()))
        yang_ming_bookings = list(map(str, df[df["Carrier Name"] == "Yang Ming Line"]["Booking Number"].tolist()))
        maersk_bookings = list(map(str, df[df["Carrier Name"] == "Maersk Line"]["Booking Number"].tolist()))
        one_bookings = list(map(str, df[df["Carrier Name"] == "Ocean Network Express"]["Booking Number"].tolist()))
        #oocl_bookings = df[df["Carrier Name"] == "OOCL"]["Booking Number"].tolist()
        print(f"Evergreen length:{len(evergreen_bookings)}")
        print(f"Wan Hai length:{len(wan_hai_bookings)}")
        print(f"Yang Ming length:{len(yang_ming_bookings)}")
        print(f"Maersk length:{len(maersk_bookings)}")
        print(f"ONE length:{len(one_bookings)}")
        
        #以下全部換成play和API後可能不需要
        # Get the directory of the current script or the temporary directory when running as an .exe
        # Determine the base directory
        # if getattr(sys, 'frozen', False):  # Check if running as a PyInstaller bundle
        #     print("run as bundle")
        #     base_dir = os.path.dirname(sys.executable) #base_dir = sys._MEIPASS 會在\dist\Main\_internal，如果打包成onefile裡面有Chromdriver的話可以用，因為driver在"裡面"
        # else:
        #     print("run as script")
        #     base_dir = os.path.dirname(os.path.abspath(__file__))  # For --onedir or running as a script

        # # Call each functions once with all booking numbers
        if len(evergreen_bookings) > 0 and "Evergreen" in CARRIER_LIST:
            print("evergreen search")
            evergreen_results = Evergreen_play.Evergreen(evergreen_bookings)
            #evergreen_results = []
        else:
            evergreen_results = []

        if len(wan_hai_bookings) > 0 and "Wan Hai" in CARRIER_LIST:
            print("wanhai search")
            wan_hai_results = asyncio.run(WanHai_play.WanHai(wan_hai_bookings))
            #wan_hai_results = WanHai.WanHai(wan_hai_bookings,driver)
        else:
            wan_hai_results = []

        if len(yang_ming_bookings) > 0 and "Yang Ming Line" in CARRIER_LIST:
            print("yangming search")
            yang_ming_results = YangMing_API.YangMing(yang_ming_bookings)
            #yang_ming_results = YangMing.YangMing(yang_ming_bookings,driver)
        else:
            yang_ming_results = []
        
        if len(maersk_bookings) > 0 and "Maersk Line" in CARRIER_LIST:
            print("maersk search")
            maersk_results = Maersk_API.Maersk(maersk_bookings)
            #maersk_results = Maersk.Maersk(maersk_bookings,driver)
        else:
            maersk_results = []
        
        if len(one_bookings) > 0 and "Ocean Network Express" in CARRIER_LIST:
            print("one search")
            one_results = ONE_API.ONE(one_bookings)
            #one_results = ONE.ONE(one_bookings,driver)
        else:
            one_results = []

        #把dicts併起來做成dataframe
        # Prepare the final DataFrame        
        final_data = pd.concat([pd.DataFrame(lst) for lst in [evergreen_results,wan_hai_results,yang_ming_results,maersk_results,one_results]], ignore_index=True)
        #final_data = pd.concat([pd.DataFrame(lst) for lst in [one_results]], ignore_index=True)
        print(final_data.to_string())
        
        track_end = time.time()
        print(f"Total tracking execution time: {track_end - track_start:.1f} seconds")
        print(f"actual search length: {len(final_data)}")
        final_data.to_csv(RESULT_DF_OUTPUT_PATH, encoding='utf-8', index=False)
        print(f"Data saved to '{RESULT_DF_OUTPUT_PATH}'")

        return final_data
    
    except Exception as e:
        print("Error", f"Failed to track: {e}")




#####------------------------Main function-------------------------------########
def tracker_main(po_update_path):
    print("Starting tracker module...")
    program_start = time.time()

    print(os.getcwd())

    sheet_name = "大表"
    #recalculate 重算一次讓值全部進入快取裡，待會才能使用Cargoo公式產出的資料，目前使用只有值的暫存檔，已不需要
    #support_function.recalculate_workbook(po_update_path)

    df = pd.read_excel(po_update_path, sheet_name=sheet_name,header=1)  # Column names are in the second row
    print(f"total length: {len(df)}") #3114

    # Filter rows where "Carrier" and "Booking" are not null, and save to a new CSV for reference
    filtered_df = df[(df["Carrier"].notna()) & (df["Booking"].notna())]
    print(f"filtered df: {len(filtered_df)}") #421
    #print(filtered_df.to_string().encode('utf-8', errors='replace').decode('utf-8'))
    filtered_df.to_csv(FILTERED_DF_OUTPUT_PATH, encoding='utf-8', index=False)
    print(f"Data saved to '{FILTERED_DF_OUTPUT_PATH}'")

    # Extract "Carrier" and "Booking" columns
    carrier_booking_df = filtered_df[["Carrier", "Booking"]]
    carrier_booking_df = carrier_booking_df.drop_duplicates()

    # Apply the track function
    data = pd.DataFrame(carrier_booking_df[["Carrier","Booking"]]) #正式測試要改為全部!!![0:3]
    print(f"total search length(all carriers): {len(data)}") #172
    #print(data.to_string()) #檢查所有Carrier和Booking資訊
    result_df = track(data)
    if result_df is None or result_df.empty:
        result_df = pd.DataFrame(columns=["booking No", "carrier", "Vessel", "ETD", "ETA", "CYtime"])


    #####------------------------開寫-------------------------------######
    # Open workbook for writing (preserve formulas) and a second read-only view
    wb = load_workbook(po_update_path)
    ws = wb[sheet_name]

    # Open a second workbook with data_only=True to get Excel's cached/calculated values
    # Use this sheet for reading runtime values so we can preserve formulas when saving
    wb_values = load_workbook(po_update_path, data_only=True)
    ws_values = wb_values[sheet_name]

    # Map column names to their respective column letters in the Excel file
    header_row = 2  # Assuming column names are in the second row
    columns = {cell.value: cell.column for cell in ws[header_row]}

    carrier_col = columns["Carrier"]
    booking_col = columns["Booking"]
    carrier_etd_col = columns["Carrier ETD"]
    carrier_eta_col = columns["Carrier ETA"]
    vessel_col = columns["vessel"]
    CYtime_col = columns["抵達櫃場時間"]
    qty_col = columns["Still to be delivered (qty)"]


    # Iterate through the rows in the Excel file and update the relevant columns
    # Read values from ws_values (data_only) so formula cells return their cached values.
    for row in ws.iter_rows(min_row=3, max_row=len(df)+2):  # Data starts from row 3，正式測試要改為max_row=len(df)+2
        
        try: 
            if row[qty_col - 1].value == None: #看是不是最後一列
                break

            # Get the row number
            row_number = row[0].row
            # Prefer the cached/calculated value from ws_values; fall back to the original cell if missing
            carrier = ws_values.cell(row=row_number, column=carrier_col).value
            if carrier is None:
                carrier = "carrier empty"
            booking = ws_values.cell(row=row_number, column=booking_col).value
            if booking is None:
                booking = "booking empty"   #row[booking_col - 1].value

            #print(f"Row {row_number}: carrier={carrier}, booking={booking}")

            # Find the matching row in result_df
            match = result_df[(result_df["carrier"] == carrier) & (result_df["booking No"] == booking)]

            #print("match ended, go to")
            if not match.empty:
                # Update the columns in the Excel file
                ws.cell(row=row[0].row, column=carrier_etd_col, value=match["ETD"].values[0])
                ws.cell(row=row[0].row, column=carrier_eta_col, value=match["ETA"].values[0])
                ws.cell(row=row[0].row, column=vessel_col, value=match["Vessel"].values[0])
                ws.cell(row=row[0].row, column=CYtime_col, value=match["CYtime"].values[0])
            elif (carrier != "carrier empty") and (carrier not in CARRIER_LIST):
                ws.cell(row=row[0].row, column=carrier_etd_col, value="carrier not exists")
                ws.cell(row=row[0].row, column=carrier_eta_col, value="carrier not exists")
                ws.cell(row=row[0].row, column=vessel_col, value="carrier not exists")
                ws.cell(row=row[0].row, column=CYtime_col, value="carrier not exists")
            elif (carrier != "carrier empty") and (carrier in CARRIER_LIST): #不應該發生 
                ws.cell(row=row[0].row, column=carrier_etd_col, value="not searched")
                ws.cell(row=row[0].row, column=carrier_eta_col, value="not searched")
                ws.cell(row=row[0].row, column=vessel_col, value="not searched")
                ws.cell(row=row[0].row, column=CYtime_col, value="not searched")
            else: #cargoo尚未提供carrier
                ws.cell(row=row[0].row, column=carrier_etd_col, value="-")
                ws.cell(row=row[0].row, column=carrier_eta_col, value="-")
                ws.cell(row=row[0].row, column=vessel_col, value="-")
                ws.cell(row=row[0].row, column=CYtime_col, value="-")
        except Exception as e:
                print(e)
                ws.cell(row=row[0].row, column=carrier_etd_col, value="fill in error")
                ws.cell(row=row[0].row, column=carrier_eta_col, value="fill in error")
                ws.cell(row=row[0].row, column=vessel_col, value="fill in error")
                ws.cell(row=row[0].row, column=CYtime_col, value="fill in error")        

    # Save the workbook
    wb.save(po_update_path)
    #實測可以保留公式W

    program_end = time.time()
    print(f"Total program execution time: {program_end - program_start:.1f} seconds")
    print("Tracker module finished, results saved to Excel.")
    print("-" * 50)


if __name__ == "__main__":
    file = support_function.get_latest_file(r"YOUR_NETWORK_DRIVE\ShipmentTracking\PO update\temp")
    tracker_main(file)